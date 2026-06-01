"""
Fiscal-period revenue report — Excel export.

GET /api/revenue_export?period=6&year=2026
    Returns a two-sheet .xlsx workbook:
      Sheet 1 "Product Mix"  — GoTab / Events / Combined breakdown by category
      Sheet 2 "Forecast"     — Food vs FOH: prior-year same-fiscal-day forecast,
                               current-period actuals, and variance

Fiscal calendar (FY2026 starts Monday 2025-12-29):
    Periods 1, 4, 7, 10  → 5 weeks    |    All other periods → 4 weeks
Prior year offset: 364 days (52 weeks) — gives the same fiscal day of week + week position
"""

import io
import json
import urllib.request
from datetime import date, timedelta
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
CRON_SECRET  = os.environ.get("CRON_SECRET", "")

# ---------------------------------------------------------------------------
# Revenue columns (must match daily_revenue table and revenue_report.py)
# ---------------------------------------------------------------------------
COLS = [
    "food", "beverage", "mini_golf", "bowling", "karaoke", "darts",
    "shuffle_board", "pool", "merchandise", "events", "open_item",
    "bottle_svc", "reservations", "gift_card", "redeemed",
]
COL_LABELS = [
    "FOOD", "BEVERAGE", "MINI GOLF", "BOWLING", "KARAOKE", "DARTS",
    "SHUFFLE BOARD", "POOL", "MERCHANDISE", "EVENTS", "OPEN ITEM",
    "BOTTLE SVC", "RESERVATIONS", "GIFT CARD", "REDEEMED",
]

DATA_START_COL = 3
TOTAL_COL      = DATA_START_COL + len(COLS)
NUM_DATA_COLS  = len(COLS)

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# GoTab categories that map to "food" — used by forecast sheet
_FORECAST_FOOD_CATS = frozenset({
    "chicken.", "dessert", "event food", "extra sauces and cheese dips",
    "fry platters", "half pound burgers", "legacy menu items",
    "pizza and flatbreads", "pretzels", "tacos", "tater kegs", "wraps",
})

# ---------------------------------------------------------------------------
# Fiscal calendar
# ---------------------------------------------------------------------------
FY_START = date(2025, 12, 29)           # P1 W1 Mon, FY2026
PRIOR_YEAR_OFFSET = timedelta(days=364) # 52 weeks — same fiscal day in prior year

_WEEKS_PER_PERIOD = {
    1: 5, 2: 4, 3: 4,
    4: 5, 5: 4, 6: 4,
    7: 5, 8: 4, 9: 4,
    10: 5, 11: 4, 12: 4,
}


def period_date_range(period: int, year: int = 2026):
    fy_offset_years = year - 2026
    fy_start = FY_START + timedelta(weeks=52 * fy_offset_years)
    weeks_before = sum(_WEEKS_PER_PERIOD[p] for p in range(1, period))
    start     = fy_start + timedelta(weeks=weeks_before)
    num_weeks = _WEEKS_PER_PERIOD[period]
    end       = start + timedelta(weeks=num_weeks) - timedelta(days=1)
    return start, end, num_weeks


def date_to_period(d: date) -> tuple:
    offset_days = (d - FY_START).days
    if offset_days < 0 or offset_days >= 52 * 7:
        raise ValueError(f"Date {d} is outside FY2026")
    offset_weeks = offset_days // 7
    total = 0
    for p in range(1, 13):
        total += _WEEKS_PER_PERIOD[p]
        if offset_weeks < total:
            return p, 2026
    raise ValueError(f"Date {d} is outside FY2026")


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------
def _supa_get(path: str) -> list:
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1{path}",
        headers={
            "apikey":        SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Accept":        "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


def _fetch_revenue(start: date, end: date) -> list:
    """Fetch daily_revenue rows for current period (product mix sheet source).
    Returns empty list if the table doesn't exist yet (pre-migration)."""
    path = (
        f"/daily_revenue"
        f"?report_date=gte.{start.isoformat()}"
        f"&report_date=lte.{end.isoformat()}"
        f"&order=report_date.asc,source.asc"
        f"&select=report_date,source,{','.join(COLS)}"
    )
    try:
        return _supa_get(path)
    except Exception:
        return []


def _build_data_map(rows: list) -> dict:
    empty = lambda: {c: 0.0 for c in COLS}
    data  = {}
    for row in rows:
        d = row["report_date"]
        if d not in data:
            data[d] = {"gotab": empty(), "tripleseat": empty()}
        src = row["source"]
        for c in COLS:
            data[d][src][c] = float(row.get(c) or 0)
    return data


def _supa_get_paged(base_path: str) -> list:
    """Paginate a Supabase query in 1000-row chunks until exhausted."""
    all_rows = []
    offset   = 0
    while True:
        sep  = "&" if "?" in base_path else "?"
        rows = _supa_get(f"{base_path}{sep}limit=1000&offset={offset}")
        all_rows.extend(rows)
        if len(rows) < 1000:
            break
        offset += 1000
    return all_rows


def _fetch_food_foh(start: date, end: date) -> dict:
    """
    Compute (food, foh, total) for each date in a range directly from
    source tables (sales + ts_events).  Used for the forecast sheet so
    it works even before daily_revenue is backfilled.
    """
    gt_rows = _supa_get_paged(
        f"/sales"
        f"?report_date=gte.{start.isoformat()}"
        f"&report_date=lte.{end.isoformat()}"
        f"&select=report_date,category,net_sales"
    )
    ts_rows = _supa_get_paged(
        f"/ts_events"
        f"?event_date=gte.{start.isoformat()}"
        f"&event_date=lte.{end.isoformat()}"
        f"&deleted_at=is.null"
        f"&select=event_date,food_amount,actual_amount"
    )

    buckets = {}  # date_str -> [food_sum, total_sum]
    for row in gt_rows:
        d   = row["report_date"]
        val = float(row.get("net_sales") or 0)
        if d not in buckets:
            buckets[d] = [0.0, 0.0]
        buckets[d][1] += val
        if (row.get("category") or "").lower().strip() in _FORECAST_FOOD_CATS:
            buckets[d][0] += val

    for row in ts_rows:
        d     = row["event_date"]
        food  = float(row.get("food_amount")  or 0)
        total = float(row.get("actual_amount") or 0)
        if d not in buckets:
            buckets[d] = [0.0, 0.0]
        buckets[d][0] += food
        buckets[d][1] += total

    result = {}
    for d, (food, total) in buckets.items():
        food  = round(food,  2)
        total = round(total, 2)
        result[d] = (food, round(total - food, 2), total)
    return result


def _day_food_foh(d_str: str, data_map: dict) -> tuple:
    """Derive (food, foh, total) from a data_map entry (product mix data)."""
    entry = data_map.get(d_str, {})
    gt    = entry.get("gotab",      {c: 0.0 for c in COLS})
    ts    = entry.get("tripleseat", {c: 0.0 for c in COLS})
    food  = (gt.get("food") or 0) + (ts.get("food") or 0)
    total = sum((gt.get(c) or 0) + (ts.get(c) or 0) for c in COLS)
    return round(food, 2), round(total - food, 2), round(total, 2)


# ---------------------------------------------------------------------------
# Shared Excel styling
# ---------------------------------------------------------------------------
_CURRENCY_FMT = '"$"#,##0.00'
_PCT_FMT      = '0.0%'


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _set_h(ws, row: int, h: float):
    ws.row_dimensions[row].height = h


# ---------------------------------------------------------------------------
# Product Mix sheet helpers (Sheet 1)
# ---------------------------------------------------------------------------
def _pm_section_header(ws, row: int, label: str, color: str) -> int:
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COL)
    _set_h(ws, row, 20)
    return row + 1


def _pm_col_headers(ws, row: int, color: str) -> int:
    for col in range(1, TOTAL_COL + 1):
        ws.cell(row=row, column=col).fill = _fill(color)
    for i, label in enumerate(COL_LABELS):
        c = ws.cell(row=row, column=DATA_START_COL + i, value=label)
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c = ws.cell(row=row, column=TOTAL_COL, value="TOTAL")
    c.font      = Font(bold=True, size=9, color="FFFFFF")
    c.fill      = _fill(color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    _set_h(ws, row, 28)
    return row + 1


def _pm_data_row(ws, row: int, day_name: str, dt: date, amounts: dict) -> int:
    ws.cell(row=row, column=1, value=day_name).font = Font(size=10)
    dc = ws.cell(row=row, column=2, value=dt)
    dc.number_format = "M/D/YYYY"
    dc.font          = Font(size=10)
    total = 0.0
    for i, col in enumerate(COLS):
        val = float(amounts.get(col) or 0)
        c   = ws.cell(row=row, column=DATA_START_COL + i, value=val)
        c.number_format = _CURRENCY_FMT
        c.font          = Font(size=10)
        c.alignment     = Alignment(horizontal="right")
        total += val
    t = ws.cell(row=row, column=TOTAL_COL, value=round(total, 2))
    t.number_format = _CURRENCY_FMT
    t.font          = Font(size=10, bold=True)
    t.alignment     = Alignment(horizontal="right")
    _set_h(ws, row, 16)
    return row + 1


def _pm_subtotal_row(ws, row: int, week_label: str, sums: dict, color: str) -> int:
    ws.cell(row=row, column=1, value="")
    lc = ws.cell(row=row, column=2, value=week_label)
    lc.font = Font(bold=True, size=10)
    lc.fill = _fill(color)
    lc.alignment = Alignment(horizontal="right")
    total = 0.0
    for i, col in enumerate(COLS):
        val = round(sums.get(col) or 0, 2)
        c   = ws.cell(row=row, column=DATA_START_COL + i, value=val)
        c.number_format = _CURRENCY_FMT
        c.font          = Font(bold=True, size=10)
        c.fill          = _fill(color)
        c.alignment     = Alignment(horizontal="right")
        total += val
    t = ws.cell(row=row, column=TOTAL_COL, value=round(total, 2))
    t.number_format = _CURRENCY_FMT
    t.font          = Font(bold=True, size=10)
    t.fill          = _fill(color)
    t.alignment     = Alignment(horizontal="right")
    _set_h(ws, row, 16)
    return row + 1


def _pm_period_totals(ws, row: int, sums: dict, color: str) -> int:
    lc = ws.cell(row=row, column=1, value="PERIOD TOTALS")
    lc.font = Font(bold=True, size=10, color="FFFFFF")
    lc.fill = _fill(color)
    lc.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=2).fill = _fill(color)
    total = 0.0
    for i, col in enumerate(COLS):
        val = round(sums.get(col) or 0, 2)
        c   = ws.cell(row=row, column=DATA_START_COL + i, value=val)
        c.number_format = _CURRENCY_FMT
        c.font          = Font(bold=True, size=10, color="FFFFFF")
        c.fill          = _fill(color)
        c.alignment     = Alignment(horizontal="right")
        total += val
    t = ws.cell(row=row, column=TOTAL_COL, value=round(total, 2))
    t.number_format = _CURRENCY_FMT
    t.font          = Font(bold=True, size=10, color="FFFFFF")
    t.fill          = _fill(color)
    t.alignment     = Alignment(horizontal="right")
    _set_h(ws, row, 18)

    pct = row + 1
    ws.cell(pct, 1, "% of total sales").font = Font(italic=True, size=9)
    for i, col in enumerate(COLS):
        pct_val = (sums.get(col) or 0) / total if total else 0.0
        c = ws.cell(pct, DATA_START_COL + i, value=pct_val)
        c.number_format = _PCT_FMT
        c.font          = Font(italic=True, size=9)
        c.alignment     = Alignment(horizontal="right")
    _set_h(ws, pct, 14)
    return pct + 1


def _pm_write_section(ws, start_row: int, section_label: str, source: str,
                      period_dates: list, data_map: dict,
                      fill_color: str, sub_color: str) -> int:
    row = _pm_section_header(ws, start_row, section_label, fill_color)
    row = _pm_col_headers(ws, row, fill_color)

    empty       = {c: 0.0 for c in COLS}
    period_sums = {c: 0.0 for c in COLS}

    for w_idx, week in enumerate(period_dates):
        week_sums = {c: 0.0 for c in COLS}
        for d_idx, dt in enumerate(week):
            d_str = dt.isoformat()
            entry = data_map.get(d_str, {})
            if source == "combined":
                gt  = entry.get("gotab",      empty)
                ts  = entry.get("tripleseat", empty)
                amt = {c: (gt.get(c) or 0) + (ts.get(c) or 0) for c in COLS}
            else:
                amt = entry.get(source, empty)
            row = _pm_data_row(ws, row, DAY_NAMES[d_idx], dt, amt)
            for c in COLS:
                week_sums[c]   += amt.get(c) or 0
                period_sums[c] += amt.get(c) or 0

        lbl = f"Week {w_idx + 1}" if source == "tripleseat" else ""
        row = _pm_subtotal_row(ws, row, lbl, week_sums, sub_color)
        if w_idx < len(period_dates) - 1:
            _set_h(ws, row, 8)
            row += 1

    row += 1
    row = _pm_period_totals(ws, row, period_sums, fill_color)
    return row + 2


# ---------------------------------------------------------------------------
# Forecast sheet helpers (Sheet 2)
# ---------------------------------------------------------------------------
# Column layout:
#   A=Date  B=Day  C=F.Food  D=F.FOH  E=F.Total  F=sep
#   G=A.Food  H=A.FOH  I=A.Total  J=sep
#   K=V.Food  L=V.FOH  M=V.Total

_FC  = {"food": 3, "foh": 4, "total": 5}   # Forecast cols
_AC  = {"food": 7, "foh": 8, "total": 9}   # Actual cols
_VC  = {"food": 11, "foh": 12, "total": 13} # Variance cols

_BLUE_HDR  = "1F3864"
_GREEN_HDR = "1E5631"
_RED_HDR   = "7B2C2C"
_SUB_BLUE  = "2E5090"
_SUB_GREEN = "2D7A4A"
_SUB_RED   = "B05050"


def _fc_headers(ws) -> int:
    """Write the two header rows for the forecast sheet, return next data row."""
    row = 1
    # Group labels row
    for col, label, color in [
        (_FC["food"], "FORECAST",  _BLUE_HDR),
        (_AC["food"], "ACTUAL",    _GREEN_HDR),
        (_VC["food"], "VARIANCE",  _RED_HDR),
    ]:
        c = ws.cell(row=row, column=col, value=label)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    _set_h(ws, row, 20)
    row += 1

    # Sub-labels row
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="")
    for group_col, color in [(_FC["food"], _BLUE_HDR), (_AC["food"], _GREEN_HDR), (_VC["food"], _RED_HDR)]:
        for offset, label in enumerate(["Food Sales", "FOH Sales", "Total Sales"]):
            c = ws.cell(row=row, column=group_col + offset, value=label)
            c.font      = Font(bold=True, size=9, color="FFFFFF")
            c.fill      = _fill(color)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _set_h(ws, row, 24)
    return row + 1


def _fc_data_row(ws, row: int, dt: date, day_name: str,
                 f_tup: tuple, a_tup: tuple) -> int:
    dc = ws.cell(row=row, column=1, value=dt)
    dc.number_format = "M/D/YYYY"
    dc.font          = Font(size=10)
    ws.cell(row=row, column=2, value=day_name).font = Font(size=10)

    v_tup = (a_tup[0] - f_tup[0], a_tup[1] - f_tup[1], a_tup[2] - f_tup[2])

    for group_cols, values in [(_FC, f_tup), (_AC, a_tup), (_VC, v_tup)]:
        for key, col in group_cols.items():
            val = values[["food", "foh", "total"].index(key)]
            c   = ws.cell(row=row, column=col, value=round(val, 2))
            c.number_format = _CURRENCY_FMT
            c.font          = Font(size=10)
            c.alignment     = Alignment(horizontal="right")

    _set_h(ws, row, 16)
    return row + 1


def _fc_subtotal_row(ws, row: int, f_sums: list, a_sums: list) -> int:
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="")
    v_sums = [a - f for a, f in zip(a_sums, f_sums)]

    for group_cols, values, color in [
        (_FC, f_sums, _SUB_BLUE),
        (_AC, a_sums, _SUB_GREEN),
        (_VC, v_sums, _SUB_RED),
    ]:
        for key, col in group_cols.items():
            val = values[["food", "foh", "total"].index(key)]
            c   = ws.cell(row=row, column=col, value=round(val, 2))
            c.number_format = _CURRENCY_FMT
            c.font          = Font(bold=True, size=10)
            c.fill          = _fill(color)
            c.alignment     = Alignment(horizontal="right")

    _set_h(ws, row, 16)
    return row + 1


def _fc_period_totals(ws, row: int, f_sums: list, a_sums: list) -> int:
    v_sums = [a - f for a, f in zip(a_sums, f_sums)]

    lc = ws.cell(row=row, column=1, value="Period Total")
    lc.font      = Font(bold=True, size=10, color="FFFFFF")
    lc.fill      = _fill(_BLUE_HDR)
    lc.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=2, value="").fill = _fill(_BLUE_HDR)

    for group_cols, values, color in [
        (_FC, f_sums, _BLUE_HDR),
        (_AC, a_sums, _GREEN_HDR),
        (_VC, v_sums, _RED_HDR),
    ]:
        for key, col in group_cols.items():
            val = values[["food", "foh", "total"].index(key)]
            c   = ws.cell(row=row, column=col, value=round(val, 2))
            c.number_format = _CURRENCY_FMT
            c.font          = Font(bold=True, size=10, color="FFFFFF")
            c.fill          = _fill(color)
            c.alignment     = Alignment(horizontal="right")

    _set_h(ws, row, 18)
    return row + 1


def _write_forecast_sheet(wb, period: int, year: int,
                          data_map: dict, prior_data: dict,
                          period_dates: list) -> None:
    ws = wb.create_sheet(title=f"P{period} Forecast")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for col in ["C", "D", "E", "G", "H", "I", "K", "L", "M"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["F"].width = 2
    ws.column_dimensions["J"].width = 2
    ws.freeze_panes = "C3"

    row      = _fc_headers(ws)
    period_f = [0.0, 0.0, 0.0]
    period_a = [0.0, 0.0, 0.0]

    for w_idx, week in enumerate(period_dates):
        week_f = [0.0, 0.0, 0.0]
        week_a = [0.0, 0.0, 0.0]

        for d_idx, dt in enumerate(week):
            d_str      = dt.isoformat()
            prior_d    = (dt - PRIOR_YEAR_OFFSET).isoformat()

            f_tup = prior_data.get(prior_d, (0.0, 0.0, 0.0))
            a_tup = _day_food_foh(d_str, data_map)

            row = _fc_data_row(ws, row, dt, DAY_NAMES[d_idx], f_tup, a_tup)
            for i in range(3):
                week_f[i] += f_tup[i]
                week_a[i] += a_tup[i]

        row = _fc_subtotal_row(ws, row, week_f, week_a)

        # blank separator between weeks (not after last)
        if w_idx < len(period_dates) - 1:
            _set_h(ws, row, 8)
            row += 1

        for i in range(3):
            period_f[i] += week_f[i]
            period_a[i] += week_a[i]

    row += 1
    _fc_period_totals(ws, row, period_f, period_a)


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
def build_workbook(period: int, year: int, data_map: dict,
                   prior_data: dict, period_dates: list,
                   num_weeks: int) -> bytes:
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Product Mix ----
    ws1 = wb.active
    ws1.title = f"P{period} Product Mix"

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 10
    for i in range(NUM_DATA_COLS):
        ws1.column_dimensions[get_column_letter(DATA_START_COL + i)].width = 11
    ws1.column_dimensions[get_column_letter(TOTAL_COL)].width = 13
    ws1.freeze_panes = "C3"

    row = 1
    row = _pm_write_section(ws1, row, "GO TAB SALES", "gotab",
                             period_dates, data_map, "1F3864", "2E5090")
    row = _pm_write_section(ws1, row, "EVENTS", "tripleseat",
                             period_dates, data_map, "1E5631", "2D7A4A")
    row = _pm_write_section(ws1, row, "COMBINED", "combined",
                             period_dates, data_map, "3B3B3B", "5C5C5C")

    # ---- Sheet 2: Forecast ----
    _write_forecast_sheet(wb, period, year, data_map, prior_data, period_dates)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run(period: int, year: int) -> bytes:
    start, end, num_weeks = period_date_range(period, year)

    period_dates = [
        [start + timedelta(weeks=w, days=d) for d in range(7)]
        for w in range(num_weeks)
    ]

    # Current period — product mix sheet uses daily_revenue
    rows     = _fetch_revenue(start, end)
    data_map = _build_data_map(rows)

    # Prior year dates for the same fiscal period
    prior_start = start - PRIOR_YEAR_OFFSET
    prior_end   = end   - PRIOR_YEAR_OFFSET
    prior_data  = _fetch_food_foh(prior_start, prior_end)

    return build_workbook(period, year, data_map, prior_data, period_dates, num_weeks)


# ---------------------------------------------------------------------------
# Vercel handler
# ---------------------------------------------------------------------------
class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        auth = self.headers.get("authorization", "")
        if CRON_SECRET and auth != f"Bearer {CRON_SECRET}":
            self.send_response(401)
            self.end_headers()
            self.wfile.write(b"Unauthorized")
            return

        if openpyxl is None:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": "openpyxl not installed"}).encode())
            return

        qs = parse_qs(urlparse(self.path).query)
        try:
            if "date" in qs:
                from datetime import date as _d
                d = _d.fromisoformat(qs["date"][0])
                period, year = date_to_period(d)
            else:
                period = int(qs.get("period", ["6"])[0])
                year   = int(qs.get("year",   ["2026"])[0])

            xlsx_bytes = run(period, year)
            filename   = f"revenue_P{period:02d}_FY{year}.xlsx"

            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())

    def log_message(self, format, *args):
        pass
